import openpyxl as excel
import datetime

# 曜日名の定義
weekname = ["月","火","水","木","金","土", "日"]

# 新規ワークブックを作る
wb = excel.Workbook()
ws = wb.active

# 今年の9月1日を得る --- (*1)
now = datetime.datetime.now()
tm = datetime.date(now.year, 9, 1)
# 366日分を繰り返してセルに書き込む --- (*2)
for i in range(1, 367):
    # 年月日と曜日を書き込む --- (*3)
    ws.cell(column=1, row=i, value=tm.year)
    ws.cell(column=2, row=i, value=tm.month)
    ws.cell(column=3, row=i, value=tm.day)
    ws.cell(column=4, row=i, value=weekname[tm.weekday()])
    # 翌日を得る --- (*4)
    tm = tm + datetime.timedelta(days=1)

wb.save("cal.xlsx")
