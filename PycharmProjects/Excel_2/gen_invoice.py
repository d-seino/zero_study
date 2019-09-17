import openpyxl, pprint
from datetime import datetime

# ファイル名の指定など --- (*1)
file_list = "list.xlsx" # 納品物一覧
file_invoice= "【御請求書】ベース.xlsx" # 請求書のテンプレート
file_delivery = "【御見積書】ベース.xlsx" # 見積書のテンプレート
file_out_iv = "out-invoice.xlsx" # 生成する請求書
file_out_ds = "out-delivery_slip.xlsx" # 生成する見積書

# 納品物一覧を読み込む --- (*2)
wb = openpyxl.load_workbook(file_list, data_only=True) # 数式でなく値を取り出す場合
ws = wb["sheet1"] # sheet1を選ぶ
name = ws["A1"].value # 宛名を得る
list_data = ws["A3:F10"] # 任意の範囲を取得

# 請求書と領収書のテンプレートを読む --- (*3)
wb_iv = openpyxl.load_workbook(file_invoice)
ws_iv = wb_iv.active
wb_ds = openpyxl.load_workbook(file_delivery)
ws_ds = wb_ds.active

# 宛名と日付を書き込む --- (*4)
cdate = datetime.now().strftime("%Y/%m/%d")
ws_iv["A3"].value = name
ws_iv["F2"].value = cdate
ws_ds["A3"].value = name
ws_ds["F2"].value = cdate

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=12+y+1, column=0+x+1, value=v)
    ws_ds.cell(row=12+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_out_iv)
wb_ds.save(file_out_ds)
print("ok")
